import openpyxl
wb = openpyxl.load_workbook('Pagi_1_C.xlsx')
wb.active = 0 #mengaktifkan indeks ke 0
sh0 = wb.active

import openpyxl
wb = openpyxl.load_workbook('Pagi_1_C.xlsx')
wb.active = 4 #mengaktifkan indeks ke 4
sh4 = wb.active

import openpyxl
wb = openpyxl.load_workbook('Pagi_1_C.xlsx')
wb.active = 6 #mengaktifkan indeks ke 6
sh6 = wb.active

w=[] #list untuk kolom B
x=[] #list untuk kolom J/penutupan (indeks 0)
y=[] #list untuk kolom J/penutupan (indeks 4)
z=[] #list untuk kolom J/penutupan (indeks 6)

for i in range(183,202):
    w.append(sh0['B' + str(i)].value)
    x.append(sh0['J' + str(i)].value)
    y.append(sh4['J' + str(i)].value)
    z.append(sh6['J' + str(i)].value)

import matplotlib.pyplot as plt
plt.plot(w,x,color='red', marker='o')
plt.plot(w,y,color='blue', marker='o')
plt.plot(w,z,color='green', marker='o')
plt.title('Kolom Penutupan pada Perusahaan')
plt.xlabel('kode saham')
plt.ylabel('saham penutupan perusahaan')
plt.legend(['penutupan (indeks 0)','penutupan(indeks 4)','penutupan(indeks 6)'])
plt.grid(True)
plt.show()